"""
Excel helpers using openpyxl.

Pure Python Excel manipulation. Automatically installs openpyxl via micropip in Pyodide.

Usage:
    from document.xlsx_helpers import xlsx_read_cells, xlsx_write_cells
    # openpyxl is automatically installed if not present
"""

import io
from typing import Any

# Import from package __init__ (works when loaded into VFS at /home/pyodide/document/)
try:
    from . import ensure_package
except ImportError:
    # Fallback for when called from document.xlsx_helpers directly in Pyodide
    from document import ensure_package


def xlsx_get_sheet_names(xlsx_bytes: bytes) -> list[str]:
    """
    Get list of sheet names from Excel workbook.

    Args:
        xlsx_bytes: Excel file as bytes

    Returns:
        List of sheet names

    Example:
        >>> from document.xlsx_helpers import xlsx_get_sheet_names
        >>> xlsx_bytes = open('/tmp/workbook.xlsx', 'rb').read()
        >>> sheets = xlsx_get_sheet_names(xlsx_bytes)
        >>> print(sheets)
        ['Sheet1', 'Sheet2', 'Data']
    """
    ensure_package("openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    return wb.sheetnames


def xlsx_read_cells(xlsx_bytes: bytes, sheet_name: str, cells: list[str]) -> dict[str, Any]:
    """
    Read specific cells from an Excel sheet.

    Args:
        xlsx_bytes: Excel file as bytes
        sheet_name: Name of the sheet to read from
        cells: List of cell references (e.g., ["A1", "B5", "C10"])

    Returns:
        Dictionary mapping cell references to their values

    Example:
        >>> from document.xlsx_helpers import xlsx_read_cells
        >>> xlsx_bytes = open('/tmp/workbook.xlsx', 'rb').read()
        >>> values = xlsx_read_cells(xlsx_bytes, 'Sheet1', ['A1', 'B2', 'C3'])
        >>> print(values)
        {'A1': 'Name', 'B2': 42, 'C3': 3.14}
    """
    ensure_package("openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name]

    result = {}
    for cell_ref in cells:
        result[cell_ref] = ws[cell_ref].value

    return result


def xlsx_write_cells(xlsx_bytes: bytes, sheet_name: str, cells: dict[str, Any]) -> bytes:
    """
    Write values to specific cells in an Excel sheet.

    Args:
        xlsx_bytes: Excel file as bytes
        sheet_name: Name of the sheet to write to
        cells: Dictionary mapping cell references to values

    Returns:
        Modified Excel file as bytes

    Example:
        >>> from document.xlsx_helpers import xlsx_write_cells
        >>> xlsx_bytes = open('/tmp/workbook.xlsx', 'rb').read()
        >>> modified = xlsx_write_cells(xlsx_bytes, 'Sheet1', {'A1': 'Updated', 'B2': 100})
        >>> open('/tmp/output.xlsx', 'wb').write(modified)
    """
    ensure_package("openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb[sheet_name]

    for cell_ref, value in cells.items():
        ws[cell_ref] = value

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def xlsx_to_dict(
    xlsx_bytes: bytes, sheet_name: str, has_header: bool = True
) -> list[dict[str, Any]]:
    """
    Convert Excel sheet to list of dictionaries.

    Args:
        xlsx_bytes: Excel file as bytes
        sheet_name: Name of the sheet to read
        has_header: If True, first row is used as column names

    Returns:
        List of dictionaries representing rows

    Example:
        >>> from document.xlsx_helpers import xlsx_to_dict
        >>> xlsx_bytes = open('/tmp/workbook.xlsx', 'rb').read()
        >>> data = xlsx_to_dict(xlsx_bytes, 'Sheet1')
        >>> print(data)
        [{'Name': 'Alice', 'Age': 30}, {'Name': 'Bob', 'Age': 25}]
    """
    ensure_package("openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    if has_header:
        headers = rows[0]
        data_rows = rows[1:]
        return [dict(zip(headers, row)) for row in data_rows]
    else:
        # Use column indices as keys
        return [dict(enumerate(row)) for row in rows]


def xlsx_has_formulas(xlsx_bytes: bytes) -> dict[str, list[str]]:
    """
    Check which cells contain formulas in an Excel workbook.

    Scans all worksheets and returns a dictionary mapping sheet names
    to lists of cell references that contain formulas.

    Args:
        xlsx_bytes: Excel file as bytes

    Returns:
        Dictionary mapping sheet names to lists of cells with formulas

    Example:
        >>> from document.xlsx_formulas import xlsx_has_formulas
        >>> xlsx_bytes = open('/tmp/workbook.xlsx', 'rb').read()
        >>> formulas = xlsx_has_formulas(xlsx_bytes)
        >>> print(formulas)
        {'Sheet1': ['A3', 'B5', 'C10'], 'Sheet2': ['D2']}
    """
    ensure_package("openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    formula_cells = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells_with_formulas = []

        for row in ws.iter_rows():
            for cell in row:
                # Check if cell has a formula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    cells_with_formulas.append(cell.coordinate)

        if cells_with_formulas:
            formula_cells[sheet_name] = cells_with_formulas

    return formula_cells


def xlsx_read_with_formulas(xlsx_bytes: bytes) -> dict[str, dict[str, Any]]:
    """
    Read Excel file and return both values and formulas.

    Args:
        xlsx_bytes: Excel file as bytes

    Returns:
        Dictionary with 'values' and 'formulas' keys containing cell data

    Example:
        >>> from document.xlsx_formulas import xlsx_read_with_formulas
        >>> xlsx_bytes = open('/tmp/workbook.xlsx', 'rb').read()
        >>> data = xlsx_read_with_formulas(xlsx_bytes)
        >>> print(data['values']['Sheet1']['A1'])  # Cell value
        >>> print(data['formulas']['Sheet1']['A3'])  # Formula text
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required. Install with: await micropip.install('openpyxl')"
        ) from e

    # Load with formulas
    wb_formulas = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    # Load with cached values
    wb_values = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    result = {"values": {}, "formulas": {}}

    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        sheet_values = {}
        sheet_formulas = {}

        for row in ws_formulas.iter_rows():
            for cell in row:
                coord = cell.coordinate

                # Get value from data_only workbook
                value_cell = ws_values[coord]
                sheet_values[coord] = value_cell.value

                # Get formula if present
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    sheet_formulas[coord] = cell.value

        result["values"][sheet_name] = sheet_values
        if sheet_formulas:
            result["formulas"][sheet_name] = sheet_formulas

    return result
