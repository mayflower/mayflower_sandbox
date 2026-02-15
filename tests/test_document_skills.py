"""
Comprehensive document processing tests matching maistack skills.
Uses MayflowerSandboxBackend to execute Python code in the sandbox.

Tests all operations from the helpers/document/ package:
- xlsx_helpers.py  (Excel: openpyxl-based read/write)
- pdf_creation.py  (PDF: fpdf2-based creation)
- pdf_manipulation.py (PDF: pypdf-based merge/split/extract)
- pptx_ooxml.py    (PowerPoint: pure OOXML manipulation)
- docx_ooxml.py    (Word: pure OOXML creation/extraction)

Test Strategy:
- All assertions use deterministic VFS verification
- Execute Python scripts directly via backend.aexecute("python script.py")
- Verify file existence and format directly in database
"""

import os

import asyncpg
import pytest
from dotenv import load_dotenv

load_dotenv()

from mayflower_sandbox.deepagents_backend import MayflowerSandboxBackend


async def vfs_read_file(db_pool, thread_id: str, path: str) -> bytes | None:
    """Read file content from VFS. Returns None if file doesn't exist."""
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = $1 AND file_path = $2
            """,
            thread_id,
            path,
        )
        return row["content"] if row else None


async def vfs_file_exists(db_pool, thread_id: str, path: str) -> bool:
    """Check if file exists in VFS."""
    return await vfs_read_file(db_pool, thread_id, path) is not None


@pytest.fixture
async def db_pool():
    """Create test database connection pool."""
    pool = await asyncpg.create_pool(
        host=os.getenv("POSTGRES_HOST", "localhost"),
        database=os.getenv("POSTGRES_DB", "mayflower_test"),
        user=os.getenv("POSTGRES_USER", "postgres"),
        password=os.getenv("POSTGRES_PASSWORD", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
    )

    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_sessions (thread_id, expires_at)
            VALUES ('doc_skills', NOW() + INTERVAL '1 day')
            ON CONFLICT (thread_id) DO NOTHING
        """
        )

    yield pool
    await pool.close()


@pytest.fixture
async def backend(db_pool):
    """Create MayflowerSandboxBackend instance."""
    return MayflowerSandboxBackend(db_pool, thread_id="doc_skills", allow_net=True)


@pytest.fixture
async def clean_files(db_pool):
    """Clean files before each test."""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM sandbox_filesystem WHERE thread_id = 'doc_skills'")
    yield


# ============================================================================
# EXCEL TESTS (helpers/document/xlsx_helpers.py)
# ============================================================================


@pytest.mark.asyncio
async def test_excel_create_workbook(backend, db_pool, clean_files):
    """Test: Create Excel workbook using openpyxl (auto-installed by ensure_package)."""
    code = """
from document import ensure_package
ensure_package("openpyxl")

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sales"

# Write headers and data
ws.append(["Product", "Quantity", "Price"])
ws.append(["Laptop", 5, 1200])
ws.append(["Mouse", 25, 30])

import io
buf = io.BytesIO()
wb.save(buf)

with open("/tmp/sales.xlsx", "wb") as f:
    f.write(buf.getvalue())

print("Excel file created successfully")
"""

    await backend.aupload_files([("/tmp/create_excel.py", code.encode())])
    result = await backend.aexecute("python /tmp/create_excel.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Deterministic VFS verification - Excel files are ZIP-based (PK magic bytes)
    content = await vfs_read_file(db_pool, "doc_skills", "/tmp/sales.xlsx")
    assert content is not None, "Excel file was not created"
    assert len(content) > 100, "Excel file seems too small"
    assert content[:2] == b"PK", "Excel file should be ZIP-based (OOXML)"


@pytest.mark.asyncio
async def test_excel_read_with_helpers(backend, db_pool, clean_files):
    """Test: Create Excel then read back with xlsx_to_dict helper."""
    code = """
from document import ensure_package
ensure_package("openpyxl")

from openpyxl import Workbook

# Create workbook with data
wb = Workbook()
ws = wb.active
ws.title = "Data"
ws.append(["Name", "Score"])
ws.append(["Alice", 95])
ws.append(["Bob", 87])

import io
buf = io.BytesIO()
wb.save(buf)
xlsx_bytes = buf.getvalue()

with open("/tmp/data.xlsx", "wb") as f:
    f.write(xlsx_bytes)

# Read back using the helper
from document.xlsx_helpers import xlsx_to_dict
result = xlsx_to_dict(xlsx_bytes, "Data")
print(f"Read {len(result)} rows")
print(f"Alice's score: {result[0]['Score']}")
"""

    await backend.aupload_files([("/tmp/read_excel.py", code.encode())])
    result = await backend.aexecute("python /tmp/read_excel.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    content = await vfs_read_file(db_pool, "doc_skills", "/tmp/data.xlsx")
    assert content is not None, "Excel file was not created"
    assert content[:2] == b"PK", "Excel file should be ZIP-based (OOXML)"


# ============================================================================
# PDF TESTS (helpers/document/pdf_creation.py, pdf_manipulation.py)
# ============================================================================


@pytest.mark.asyncio
async def test_pdf_create(backend, db_pool, clean_files):
    """Test: pdf_create_simple() from document.pdf_creation."""
    code = """
from document.pdf_creation import pdf_create_simple

path = pdf_create_simple(
    "Q4 Financial Report",
    ["Revenue increased 25%", "Customer growth at 40%"],
    output_path="/tmp/report.pdf",
)

print(f"PDF created at {path}")
"""

    await backend.aupload_files([("/tmp/create_pdf.py", code.encode())])
    result = await backend.aexecute("python /tmp/create_pdf.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Deterministic VFS verification - PDF files start with %PDF
    content = await vfs_read_file(db_pool, "doc_skills", "/tmp/report.pdf")
    assert content is not None, "PDF file was not created"
    assert content[:4] == b"%PDF", "PDF file should start with %PDF header"


@pytest.mark.asyncio
async def test_pdf_merge(backend, db_pool, clean_files):
    """Test: pdf_merge() from document.pdf_manipulation."""
    code = """
from document.pdf_creation import pdf_create_simple
from document.pdf_manipulation import pdf_merge

# Create two PDFs
pdf_create_simple("Page 1", ["Content for page 1"], output_path="/tmp/doc1.pdf")
pdf_create_simple("Page 2", ["Content for page 2"], output_path="/tmp/doc2.pdf")

# Read them back
with open("/tmp/doc1.pdf", "rb") as f1, open("/tmp/doc2.pdf", "rb") as f2:
    merged = pdf_merge([f1.read(), f2.read()])

with open("/tmp/merged.pdf", "wb") as f:
    f.write(merged)

print("PDFs merged successfully")
"""

    await backend.aupload_files([("/tmp/merge_pdf.py", code.encode())])
    result = await backend.aexecute("python /tmp/merge_pdf.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Deterministic VFS verification - all three PDFs should exist
    for path in ["/tmp/doc1.pdf", "/tmp/doc2.pdf", "/tmp/merged.pdf"]:
        content = await vfs_read_file(db_pool, "doc_skills", path)
        assert content is not None, f"{path} was not created"
        assert content[:4] == b"%PDF", f"{path} should be a valid PDF"


# ============================================================================
# POWERPOINT TESTS (helpers/document/pptx_ooxml.py)
# ============================================================================


@pytest.mark.asyncio
async def test_powerpoint_create_and_extract(backend, db_pool, clean_files):
    """Test: Build a minimal PPTX using OOXML, then extract text with pptx_extract_text."""
    code = """
import io
import zipfile

# Build a minimal valid PPTX with one slide containing text.
# A PPTX is a ZIP with specific OOXML XML files.

content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/ppt/presentation.xml"
    ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>
  <Override PartName="/ppt/slides/slide1.xml"
    ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>
</Types>'''

rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
    Target="ppt/presentation.xml"/>
</Relationships>'''

presentation = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:presentation xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"
                xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <p:sldIdLst>
    <p:sldId id="256" r:id="rId2"/>
  </p:sldIdLst>
</p:presentation>'''

ppt_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId2"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide"
    Target="slides/slide1.xml"/>
</Relationships>'''

slide1 = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
       xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"
       xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <p:cSld>
    <p:spTree>
      <p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>
      <p:grpSpPr/>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="2" name="Title"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
        <p:spPr/>
        <p:txBody>
          <a:bodyPr/>
          <a:p><a:r><a:t>Company Code: XYZ789</a:t></a:r></a:p>
          <a:p><a:r><a:t>Important info</a:t></a:r></a:p>
        </p:txBody>
      </p:sp>
    </p:spTree>
  </p:cSld>
</p:sld>'''

buf = io.BytesIO()
with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
    zf.writestr("[Content_Types].xml", content_types)
    zf.writestr("_rels/.rels", rels)
    zf.writestr("ppt/presentation.xml", presentation)
    zf.writestr("ppt/_rels/presentation.xml.rels", ppt_rels)
    zf.writestr("ppt/slides/slide1.xml", slide1)

pptx_bytes = buf.getvalue()

with open("/tmp/pitch.pptx", "wb") as f:
    f.write(pptx_bytes)

# Extract text with the helper
from document.pptx_ooxml import pptx_extract_text
texts = pptx_extract_text(pptx_bytes)
print(f"Extracted from {len(texts)} slides")
for slide_num, slide_texts in texts.items():
    print(f"  Slide {slide_num}: {slide_texts}")

# Write extracted text to verify
with open("/tmp/extracted.txt", "w") as f:
    for slide_texts in texts.values():
        f.write("\\n".join(slide_texts))

print("PowerPoint created and text extracted successfully")
"""

    await backend.aupload_files([("/tmp/create_pptx.py", code.encode())])
    result = await backend.aexecute("python /tmp/create_pptx.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Verify PPTX file
    content = await vfs_read_file(db_pool, "doc_skills", "/tmp/pitch.pptx")
    assert content is not None, "PowerPoint file was not created"
    assert content[:2] == b"PK", "PowerPoint file should be ZIP-based (OOXML)"

    # Verify extraction worked
    extracted = await vfs_read_file(db_pool, "doc_skills", "/tmp/extracted.txt")
    assert extracted is not None, "Extracted text file was not created"
    assert b"XYZ789" in extracted, f"Expected company code. Got: {extracted.decode()}"


# ============================================================================
# WORD TESTS (helpers/document/docx_ooxml.py)
# ============================================================================


@pytest.mark.asyncio
async def test_word_create_document(backend, db_pool, clean_files):
    """Test: create_docx_bytes() from document.docx_ooxml."""
    code = """
from document.docx_ooxml import create_docx_bytes

paragraphs = [
    "Annual Report 2024",
    "Overview - Business performance summary",
    "Achievements:",
    "Revenue growth 35%",
    "Customer satisfaction 95%",
]

docx_bytes = create_docx_bytes(paragraphs)

with open("/tmp/report.docx", "wb") as f:
    f.write(docx_bytes)

print("Word document created successfully")
"""

    await backend.aupload_files([("/tmp/create_docx.py", code.encode())])
    result = await backend.aexecute("python /tmp/create_docx.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Deterministic VFS verification - DOCX files are ZIP-based (PK magic bytes)
    content = await vfs_read_file(db_pool, "doc_skills", "/tmp/report.docx")
    assert content is not None, "Word document was not created"
    assert content[:2] == b"PK", "Word document should be ZIP-based (OOXML)"


@pytest.mark.asyncio
async def test_word_extract_text(backend, db_pool, clean_files):
    """Test: create_docx_bytes() + docx_to_markdown() from document.docx_ooxml."""
    code = """
from document.docx_ooxml import create_docx_bytes, docx_to_markdown

paragraphs = ["Access Code: BETA456", "Additional info here"]
docx_bytes = create_docx_bytes(paragraphs)

with open("/tmp/secret.docx", "wb") as f:
    f.write(docx_bytes)

# Extract text
with open("/tmp/secret.docx", "rb") as f:
    text = docx_to_markdown(f.read())

with open("/tmp/extracted.txt", "w") as f:
    f.write(text)

print("Text extracted:", text)
"""

    await backend.aupload_files([("/tmp/extract_docx.py", code.encode())])
    result = await backend.aexecute("python /tmp/extract_docx.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Verify Word document was created with correct format
    docx_content = await vfs_read_file(db_pool, "doc_skills", "/tmp/secret.docx")
    assert docx_content is not None, "Word document was not created"
    assert docx_content[:2] == b"PK", "Word document should be ZIP-based (OOXML)"

    # Verify extraction worked
    extracted = await vfs_read_file(db_pool, "doc_skills", "/tmp/extracted.txt")
    assert extracted is not None, "Extracted text file was not created"
    assert b"BETA456" in extracted, f"Expected access code. Got: {extracted.decode()}"


# ============================================================================
# INTEGRATION TESTS (Complex workflows)
# ============================================================================


@pytest.mark.asyncio
async def test_multi_format_workflow(backend, db_pool, clean_files):
    """Test creating multiple document formats in one workflow."""
    code = """
# Excel - use openpyxl directly (auto-installed by ensure_package)
from document import ensure_package
ensure_package("openpyxl")
from openpyxl import Workbook
import io

wb = Workbook()
ws = wb.active
ws.title = "Sales"
ws.append(["Product", "Q1", "Q2"])
ws.append(["Widget", 100, 150])
ws.append(["Gadget", 200, 250])
buf = io.BytesIO()
wb.save(buf)
with open("/tmp/sales.xlsx", "wb") as f:
    f.write(buf.getvalue())

# PDF - use pdf_create_simple helper
from document.pdf_creation import pdf_create_simple
pdf_create_simple("Sales Report", ["Quarterly sales summary"], output_path="/tmp/sales.pdf")

# Word - use create_docx_bytes helper
from document.docx_ooxml import create_docx_bytes
docx = create_docx_bytes(["Executive Summary", "Q1 and Q2 sales exceeded targets."])
with open("/tmp/sales.docx", "wb") as f:
    f.write(docx)

print("All three documents created successfully")
"""

    await backend.aupload_files([("/tmp/multi_format.py", code.encode())])
    result = await backend.aexecute("python /tmp/multi_format.py")
    assert result.exit_code == 0, f"Failed: {result.output}"

    # Deterministic VFS verification - all three files should exist
    xlsx = await vfs_read_file(db_pool, "doc_skills", "/tmp/sales.xlsx")
    assert xlsx is not None, "Excel file was not created"
    assert xlsx[:2] == b"PK", "Excel file should be ZIP-based"

    pdf = await vfs_read_file(db_pool, "doc_skills", "/tmp/sales.pdf")
    assert pdf is not None, "PDF file was not created"
    assert pdf[:4] == b"%PDF", "PDF file should start with %PDF header"

    docx = await vfs_read_file(db_pool, "doc_skills", "/tmp/sales.docx")
    assert docx is not None, "Word document was not created"
    assert docx[:2] == b"PK", "Word document should be ZIP-based"
