"""
Tests for error boundaries and edge cases.

These tests verify the sandbox handles edge cases gracefully:
- Corrupt/malformed documents
- Memory limits
- Basic error handling

Note: Timeout tests are in test_timeout.py to avoid worker pool state issues.
"""

import os

import asyncpg
import pytest

from mayflower_sandbox.sandbox_executor import SandboxExecutor


@pytest.fixture
async def db_pool():
    """Create test database connection pool."""
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "database": os.getenv("POSTGRES_DB", "mayflower_test"),
        "user": os.getenv("POSTGRES_USER", "postgres"),
        "password": os.getenv("POSTGRES_PASSWORD", "postgres"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
    }

    pool = await asyncpg.create_pool(**db_config)

    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_sessions (thread_id, expires_at)
            VALUES ('test_error_boundaries', NOW() + INTERVAL '1 day')
            ON CONFLICT (thread_id) DO NOTHING
        """
        )

    yield pool
    await pool.close()


@pytest.fixture
async def clean_files(db_pool):
    """Clean files before each test."""
    async with db_pool.acquire() as conn:
        await conn.execute(
            "DELETE FROM sandbox_filesystem WHERE thread_id = 'test_error_boundaries'"
        )
    yield


class TestCorruptDocumentHandling:
    """Tests for handling corrupt/malformed documents."""

    async def test_corrupt_xlsx_handling(self, db_pool, clean_files):
        """Test handling of corrupt Excel file."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=True, timeout_seconds=60.0
        )

        code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import load_workbook
import io

# Create corrupt/invalid Excel data (not a valid ZIP)
corrupt_data = b"This is not a valid Excel file"

try:
    # Try to load corrupt file
    wb = load_workbook(io.BytesIO(corrupt_data))
    print("ERROR: Should have raised exception")
except Exception as e:
    print(f"Correctly caught error: {type(e).__name__}")
    print("Corrupt file handled gracefully")
"""

        result = await executor.execute(code)

        assert result.success is True
        assert "Corrupt file handled gracefully" in result.stdout
        assert "Correctly caught error" in result.stdout

    async def test_corrupt_docx_ooxml_handling(self, db_pool, clean_files):
        """Test handling of corrupt Word document with OOXML helpers."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=False, timeout_seconds=60.0
        )

        code = """
from document.docx_ooxml import docx_extract_text

# Test with completely invalid data
try:
    result = docx_extract_text(b"not a valid docx")
    print(f"Result: '{result}'")
    print("Invalid docx handled gracefully")
except Exception as e:
    print(f"Caught error: {type(e).__name__}")
    print("Invalid docx handled gracefully")
"""

        result = await executor.execute(code)

        # Should either return empty/error string or raise gracefully
        assert result.success is True
        assert "handled gracefully" in result.stdout

    async def test_truncated_zip_handling(self, db_pool, clean_files):
        """Test handling of truncated ZIP (incomplete download)."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=True, timeout_seconds=60.0
        )

        code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import load_workbook
import io

# Create truncated ZIP header (incomplete file)
truncated_zip = b'PK\\x03\\x04\\x14\\x00\\x00\\x00'  # ZIP magic + partial header

try:
    wb = load_workbook(io.BytesIO(truncated_zip))
    print("ERROR: Should have raised exception")
except Exception as e:
    print(f"Correctly caught: {type(e).__name__}")
    print("Truncated file handled gracefully")
"""

        result = await executor.execute(code)

        assert result.success is True
        assert "handled gracefully" in result.stdout

    async def test_malformed_xml_in_docx(self, db_pool, clean_files):
        """Test handling of valid ZIP but malformed XML content."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=False, timeout_seconds=60.0
        )

        code = """
import zipfile
import io

from document.docx_ooxml import docx_extract_text

# Create valid ZIP but with malformed XML
buf = io.BytesIO()
with zipfile.ZipFile(buf, 'w') as zf:
    # Valid content types but malformed document.xml
    zf.writestr('[Content_Types].xml', '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>')
    zf.writestr('word/document.xml', '<invalid xml with no closing tag')

docx_bytes = buf.getvalue()

try:
    text = docx_extract_text(docx_bytes)
    # Should return empty or partial result, not crash
    print(f"Extracted (may be empty): '{text}'")
    print("Malformed XML handled gracefully")
except Exception as e:
    print(f"Caught error: {type(e).__name__}")
    print("Malformed XML handled gracefully")
"""

        result = await executor.execute(code)

        assert result.success is True
        assert "handled gracefully" in result.stdout

    async def test_empty_document_handling(self, db_pool, clean_files):
        """Test handling of empty/zero-byte files."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=True, timeout_seconds=60.0
        )

        code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import load_workbook
import io

try:
    # Try to load empty file
    wb = load_workbook(io.BytesIO(b""))
    print("ERROR: Should have raised exception")
except Exception as e:
    print(f"Correctly caught: {type(e).__name__}")
    print("Empty file handled gracefully")
"""

        result = await executor.execute(code)

        assert result.success is True
        assert "handled gracefully" in result.stdout


class TestMemoryBoundaries:
    """Tests for memory limit handling."""

    async def test_large_string_creation(self, db_pool, clean_files):
        """Test creating reasonably large strings doesn't crash."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=False, timeout_seconds=30.0
        )

        code = """
# Create a moderately large string (1MB)
large_string = "x" * (1024 * 1024)
print(f"Created string of length: {len(large_string)}")
"""

        result = await executor.execute(code)

        assert result.success is True
        assert "Created string of length: 1048576" in result.stdout

    async def test_large_list_creation(self, db_pool, clean_files):
        """Test creating large lists is handled."""
        executor = SandboxExecutor(
            db_pool, "test_error_boundaries", allow_net=False, timeout_seconds=30.0
        )

        code = """
# Create a large list
large_list = list(range(100000))
print(f"Created list of length: {len(large_list)}")
print(f"Sum: {sum(large_list)}")
"""

        result = await executor.execute(code)

        assert result.success is True
        assert "Created list of length: 100000" in result.stdout
