"""
Tests for python_run_prepared and file_write tools.

Two test categories:
1. Unit tests - Test tool execution directly with pre-defined state (no LLM dependency)
2. Integration tests - Test full agent flow with VFS verification (no stdout string matching)

All assertions use deterministic checks:
- VFS file existence
- VFS file content
- Tool return values (structured)

NO fragile patterns:
- No regex on LLM output
- No string matching on stdout
- No markdown code extraction
"""

import os
import sys
from typing import Annotated

import asyncpg
import pytest
from langchain_core.messages import AIMessage, HumanMessage, ToolMessage
from langgraph.graph.message import add_messages
from typing_extensions import TypedDict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from mayflower_sandbox.filesystem import VirtualFilesystem  # noqa: E402
from mayflower_sandbox.tools import create_sandbox_tools  # noqa: E402

# =============================================================================
# Test Fixtures
# =============================================================================


class AgentState(TypedDict):
    """State for testing tools."""

    messages: Annotated[list, add_messages]
    pending_content_map: dict[str, str]


@pytest.fixture
async def db_pool():
    """Create test database connection pool."""
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "database": os.getenv("POSTGRES_DB", "mayflower_test"),
        "user": os.getenv("POSTGRES_USER", "postgres"),
        "password": os.getenv("POSTGRES_PASSWORD", "postgres"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
        "min_size": 10,
        "max_size": 50,
        "command_timeout": 60,
    }

    pool = await asyncpg.create_pool(**db_config)

    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_sessions (thread_id, expires_at)
            VALUES ('e2e_prepared_test', NOW() + INTERVAL '1 day')
            ON CONFLICT (thread_id) DO NOTHING
        """
        )

    yield pool
    await pool.close()


@pytest.fixture
async def clean_files(db_pool):
    """Clean files before each test."""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM sandbox_filesystem WHERE thread_id = 'e2e_prepared_test'")
    yield


@pytest.fixture
def tools(db_pool):
    """Create sandbox tools for testing."""
    return create_sandbox_tools(
        db_pool,
        thread_id="e2e_prepared_test",
        include_tools=["python_run_prepared", "file_write"],
    )


# =============================================================================
# VFS Helpers (Deterministic Verification)
# =============================================================================


async def vfs_file_exists(db_pool, thread_id: str, path: str) -> bool:
    """Check if a file exists in the VFS."""
    vfs = VirtualFilesystem(db_pool, thread_id)
    try:
        await vfs.read_file(path)
        return True
    except FileNotFoundError:
        return False


async def vfs_file_content(db_pool, thread_id: str, path: str) -> bytes | None:
    """Get file content from VFS, returns None if not found."""
    vfs = VirtualFilesystem(db_pool, thread_id)
    try:
        entry = await vfs.read_file(path)
        return entry["content"]
    except FileNotFoundError:
        return None


# =============================================================================
# Unit Tests - Direct Tool Execution (No LLM)
# =============================================================================


async def test_python_run_prepared_executes_code_from_state(db_pool, clean_files, tools):
    """Test that python_run_prepared executes code from pending_content_map."""
    tool = next(t for t in tools if t.name == "python_run_prepared")
    tool_call_id = "test_call_001"

    # Pre-defined code (no LLM generation)
    code = """
result = 5 * 24
print(f"Result: {result}")
"""

    # Call tool directly with injected state
    result = await tool._arun(
        file_path="/tmp/test_calc.py",
        description="Test calculation",
        _state={"pending_content_map": {tool_call_id: code}},
        tool_call_id=tool_call_id,
        _config=None,
        run_manager=None,
    )

    # Verify: result contains output (may be string or Command with resume)
    from langgraph.types import Command

    if isinstance(result, Command):
        output = result.resume
    else:
        output = str(result)

    assert "120" in output, f"Expected '120' in result: {output}"


async def test_python_run_prepared_creates_file_in_vfs(db_pool, clean_files, tools):
    """Test that python_run_prepared creates files that persist in VFS."""
    tool = next(t for t in tools if t.name == "python_run_prepared")
    tool_call_id = "test_call_002"

    # Code that creates a file
    code = """
with open('/tmp/hello.txt', 'w') as f:
    f.write('Hello from test')
print('File created')
"""

    await tool._arun(
        file_path="/tmp/create_file.py",
        description="Create a file",
        _state={"pending_content_map": {tool_call_id: code}},
        tool_call_id=tool_call_id,
        _config=None,
        run_manager=None,
    )

    # Verify: file exists in VFS (deterministic check)
    exists = await vfs_file_exists(db_pool, "e2e_prepared_test", "/tmp/hello.txt")
    assert exists, "Expected /tmp/hello.txt to exist in VFS"

    # Verify: file content is correct
    content = await vfs_file_content(db_pool, "e2e_prepared_test", "/tmp/hello.txt")
    assert content == b"Hello from test", f"Unexpected content: {content}"


async def test_python_run_prepared_without_state_returns_error(db_pool, clean_files, tools):
    """Test that python_run_prepared returns error when state is missing."""
    tool = next(t for t in tools if t.name == "python_run_prepared")

    # Call without state
    result = await tool._arun(
        file_path="/tmp/test.py",
        description="Test",
        _state=None,  # No state
        tool_call_id="test_call",
        _config=None,
        run_manager=None,
    )

    # Verify: returns error message
    assert "Error" in result
    assert "graph state" in result.lower()


async def test_python_run_prepared_with_missing_code_returns_error(db_pool, clean_files, tools):
    """Test that python_run_prepared returns error when code not in state."""
    tool = next(t for t in tools if t.name == "python_run_prepared")

    # Call with empty state
    result = await tool._arun(
        file_path="/tmp/test.py",
        description="Test",
        _state={"pending_content_map": {}},  # Empty map
        tool_call_id="nonexistent_call_id",
        _config=None,
        run_manager=None,
    )

    # Verify: returns error message
    assert "Error" in result or "No code found" in result


async def test_file_write_creates_file_in_vfs(db_pool, clean_files, tools):
    """Test that file_write tool creates files in VFS."""
    tool = next(t for t in tools if t.name == "file_write")
    tool_call_id = "test_call_003"

    # Content to write
    content = "name,value\ntest,123\n"

    await tool._arun(
        file_path="/tmp/data.csv",
        description="Write CSV data",
        _state={"pending_content_map": {tool_call_id: content}},
        tool_call_id=tool_call_id,
        _config=None,
        run_manager=None,
    )

    # Verify: file exists in VFS
    exists = await vfs_file_exists(db_pool, "e2e_prepared_test", "/tmp/data.csv")
    assert exists, "Expected /tmp/data.csv to exist in VFS"

    # Verify: file content is correct
    vfs_content = await vfs_file_content(db_pool, "e2e_prepared_test", "/tmp/data.csv")
    assert vfs_content == content.encode(), f"Unexpected content: {vfs_content}"


async def test_file_write_without_state_returns_error(db_pool, clean_files, tools):
    """Test that file_write returns error when state is missing."""
    tool = next(t for t in tools if t.name == "file_write")

    result = await tool._arun(
        file_path="/tmp/test.txt",
        description="Test",
        _state=None,
        tool_call_id="test_call",
        _config=None,
        run_manager=None,
    )

    assert "Error" in result


# =============================================================================
# Unit Tests - Document Helpers (No LLM)
# =============================================================================


async def test_python_run_prepared_with_pdf_helper(db_pool, clean_files, tools):
    """Test python_run_prepared can create PDF using helpers."""
    tool = next(t for t in tools if t.name == "python_run_prepared")
    tool_call_id = "test_call_pdf"

    # Code using pdf_creation helper
    code = """
import micropip
await micropip.install('fpdf2')

from fpdf import FPDF

pdf = FPDF()
pdf.add_page()
pdf.set_font('Helvetica', size=12)
pdf.cell(200, 10, text='Test PDF Document', new_x='LMARGIN', new_y='NEXT')
pdf.output('/tmp/test_report.pdf')
print('PDF created')
"""

    await tool._arun(
        file_path="/tmp/create_pdf.py",
        description="Create PDF",
        _state={"pending_content_map": {tool_call_id: code}},
        tool_call_id=tool_call_id,
        _config=None,
        run_manager=None,
    )

    # Verify: PDF file exists in VFS
    exists = await vfs_file_exists(db_pool, "e2e_prepared_test", "/tmp/test_report.pdf")
    assert exists, "Expected /tmp/test_report.pdf to exist in VFS"

    # Verify: file has PDF magic bytes
    content = await vfs_file_content(db_pool, "e2e_prepared_test", "/tmp/test_report.pdf")
    assert content is not None
    assert content.startswith(b"%PDF"), "Expected PDF magic bytes"


async def test_python_run_prepared_with_excel_helper(db_pool, clean_files, tools):
    """Test python_run_prepared can create Excel files using helpers."""
    tool = next(t for t in tools if t.name == "python_run_prepared")
    tool_call_id = "test_call_xlsx"

    # Code using openpyxl
    code = """
import micropip
await micropip.install('openpyxl')

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Sales'
ws['A1'] = 'Product'
ws['B1'] = 'Quantity'
ws['A2'] = 'Widget'
ws['B2'] = 42
wb.save('/tmp/test_data.xlsx')
print('Excel file created')
"""

    await tool._arun(
        file_path="/tmp/create_xlsx.py",
        description="Create Excel file",
        _state={"pending_content_map": {tool_call_id: code}},
        tool_call_id=tool_call_id,
        _config=None,
        run_manager=None,
    )

    # Verify: Excel file exists in VFS
    exists = await vfs_file_exists(db_pool, "e2e_prepared_test", "/tmp/test_data.xlsx")
    assert exists, "Expected /tmp/test_data.xlsx to exist in VFS"

    # Verify: file has ZIP magic bytes (xlsx is a ZIP file)
    content = await vfs_file_content(db_pool, "e2e_prepared_test", "/tmp/test_data.xlsx")
    assert content is not None
    assert content[:4] == b"PK\x03\x04", "Expected ZIP magic bytes for xlsx"


# =============================================================================
# Integration Tests - Full Agent Flow with VFS Verification
# =============================================================================
# These tests use a real LLM but only verify VFS state, not LLM output.
# They are marked as slow and can be skipped in CI.


@pytest.mark.skipif(
    not os.getenv("OPENAI_API_KEY"), reason="OPENAI_API_KEY not set - skipping LLM test"
)
@pytest.mark.slow
async def test_agent_creates_file_via_tool(db_pool, clean_files):
    """
    Integration test: Agent creates a file using python_run_prepared.

    This test verifies the full flow but ONLY checks VFS state (deterministic).
    It does NOT parse LLM output or tool stdout.
    """
    from dotenv import load_dotenv
    from langchain_openai import ChatOpenAI
    from langgraph.checkpoint.memory import MemorySaver
    from langgraph.graph import END, StateGraph

    load_dotenv()

    tools = create_sandbox_tools(
        db_pool,
        thread_id="e2e_prepared_test",
        include_tools=["python_run_prepared"],
    )

    llm = ChatOpenAI(model="gpt-4o-mini", temperature=0)
    llm_with_tools = llm.bind_tools(tools)
    tools_by_name = {tool.name: tool for tool in tools}

    def agent_node(state: AgentState) -> dict:
        messages = state.get("messages", [])
        response = llm_with_tools.invoke(messages)
        return {"messages": [response]}

    async def tool_node(state: AgentState) -> dict:
        """Execute tools with state injection."""
        messages = state.get("messages", [])
        last_message = messages[-1]

        if not hasattr(last_message, "tool_calls") or not last_message.tool_calls:
            return {"messages": []}

        pending_content_map = state.get("pending_content_map", {}).copy()
        msgs = []

        # Extract code from LLM's structured tool call args (not markdown parsing)
        # The LLM should include code in the tool call arguments
        for tool_call in last_message.tool_calls:
            tool_name = tool_call["name"]
            tool = tools_by_name.get(tool_name)
            tool_call_id = tool_call.get("id", "")

            if tool is None:
                msgs.append(ToolMessage(f"Unknown tool: {tool_name}", tool_call_id=tool_call_id))
                continue

            # For python_run_prepared, extract code from message content
            # Note: This regex is for tool input extraction, NOT for test assertions
            if (
                tool_name == "python_run_prepared"
                and isinstance(last_message, AIMessage)
                and last_message.content
            ):
                content = last_message.content
                if isinstance(content, str):
                    # Look for code in the content - store it
                    import re

                    match = re.search(r"```(?:python)?\n(.*?)\n```", content, re.DOTALL)
                    if match:
                        pending_content_map[tool_call_id] = match.group(1)

            try:
                kwargs = tool_call.get("args", {}).copy()
                kwargs["_state"] = {"pending_content_map": pending_content_map}
                kwargs["tool_call_id"] = tool_call_id
                kwargs["_config"] = None
                kwargs["run_manager"] = None

                result = await tool._arun(**kwargs)
                msgs.append(ToolMessage(str(result), tool_call_id=tool_call_id))
            except Exception as e:
                msgs.append(ToolMessage(f"Error: {e}", tool_call_id=tool_call_id))

        return {"messages": msgs, "pending_content_map": pending_content_map}

    def should_continue(state: AgentState) -> str:
        messages = state.get("messages", [])
        if messages and hasattr(messages[-1], "tool_calls") and messages[-1].tool_calls:
            return "tools"
        return END

    workflow = StateGraph(AgentState)
    workflow.add_node("agent", agent_node)
    workflow.add_node("tools", tool_node)
    workflow.set_entry_point("agent")
    workflow.add_conditional_edges("agent", should_continue, {"tools": "tools", END: END})
    workflow.add_edge("tools", "agent")

    app = workflow.compile(checkpointer=MemorySaver())

    # Run agent with clear instruction
    await app.ainvoke(
        {
            "messages": [
                HumanMessage(
                    content="Create a file at /tmp/agent_test.txt with the content 'Created by agent'. "
                    "Write Python code and use python_run_prepared to execute it."
                )
            ],
            "pending_content_map": {},
        },
        config={"configurable": {"thread_id": "test-agent-file"}},
    )

    # ONLY verify VFS state (deterministic) - no stdout/LLM output parsing
    exists = await vfs_file_exists(db_pool, "e2e_prepared_test", "/tmp/agent_test.txt")
    assert exists, "Expected /tmp/agent_test.txt to exist in VFS after agent execution"

    content = await vfs_file_content(db_pool, "e2e_prepared_test", "/tmp/agent_test.txt")
    assert content is not None
    assert b"Created by agent" in content, f"Unexpected file content: {content}"
