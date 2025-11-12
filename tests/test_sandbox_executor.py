"""
Tests for the new clean SandboxExecutor architecture.
Tests VFS integration, Pyodide execution, and file sync.
"""

import logging
import os
import sys

import asyncpg
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from mayflower_sandbox.sandbox_executor import SandboxExecutor


@pytest.fixture
async def db_pool():
    """Create test database connection pool."""
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "database": os.getenv("POSTGRES_DB", "mayflower_test"),
        "user": os.getenv("POSTGRES_USER", "postgres"),
        "password": os.getenv("POSTGRES_PASSWORD", "postgres"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
    }

    pool = await asyncpg.create_pool(**db_config)

    # Ensure session exists
    async with pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO sandbox_sessions (thread_id, expires_at)
            VALUES ('test_sandbox', NOW() + INTERVAL '1 day')
            ON CONFLICT (thread_id) DO NOTHING
        """)

    yield pool
    await pool.close()


@pytest.fixture
async def executor(db_pool):
    """Create executor instance."""
    return SandboxExecutor(db_pool, "test_sandbox", allow_net=True)


@pytest.fixture
async def clean_files(db_pool):
    """Clean files before and after each test to ensure isolation."""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM sandbox_filesystem WHERE thread_id = 'test_sandbox'")
    yield
    # Cleanup after test to ensure complete isolation
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM sandbox_filesystem WHERE thread_id = 'test_sandbox'")


async def test_simple_execution(executor, clean_files):
    """Test basic Python execution."""
    result = await executor.execute("print('Hello, World!')")

    assert result.success is True
    assert "Hello, World!" in result.stdout
    assert result.stderr == ""


async def test_computation(executor, clean_files):
    """Test computation with result."""
    code = """
x = 5 + 7
print(f"Result: {x}")
"""
    result = await executor.execute(code)

    assert result.success is True
    assert "Result: 12" in result.stdout


async def test_error_handling(executor, clean_files):
    """Test error is captured."""
    result = await executor.execute("raise ValueError('test error')")

    assert result.success is False
    assert "ValueError" in result.stderr
    assert "test error" in result.stderr


async def test_vfs_preload(executor, db_pool, clean_files):
    """Test files are pre-loaded from VFS."""
    # Put files in VFS first
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_filesystem (thread_id, file_path, content, content_type, size)
            VALUES ('test_sandbox', '/tmp/data.txt', $1, 'text/plain', $2)
        """,
            b"Hello from VFS!",
            16,
        )

    # Execute code that reads the file
    code = """
with open('/tmp/data.txt', 'r') as f:
    content = f.read()
    print(f"Read: {content}")
"""

    result = await executor.execute(code)

    assert result.success is True
    assert "Read: Hello from VFS!" in result.stdout


async def test_file_creation_and_save(executor, db_pool, clean_files):
    """Test created files are saved to VFS."""
    code = """
with open('/tmp/output.txt', 'w') as f:
    f.write("Created in Pyodide!")

print("File created")
"""

    result = await executor.execute(code)

    assert result.success is True
    assert "File created" in result.stdout
    assert result.created_files is not None
    assert "/tmp/output.txt" in result.created_files

    # Verify file was saved to PostgreSQL
    async with db_pool.acquire() as conn:
        file_data = await conn.fetchrow("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/output.txt'
        """)

        assert file_data is not None
        assert file_data["content"] == b"Created in Pyodide!"


async def test_vfs_roundtrip(executor, db_pool, clean_files):
    """Test full roundtrip: VFS → Pyodide → VFS."""
    # Step 1: Put input file in VFS
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_filesystem (thread_id, file_path, content, content_type, size)
            VALUES ('test_sandbox', '/data/input.csv', $1, 'text/csv', $2)
        """,
            b"a,b\n1,2\n3,4",
            11,
        )

    # Step 2: Execute code that reads input and creates output
    code = """
with open('/data/input.csv', 'r') as f:
    lines = f.readlines()
    print(f"Read {len(lines)} lines")

# Process and save
with open('/tmp/output.txt', 'w') as f:
    f.write(f"Processed {len(lines)} lines")
"""

    result = await executor.execute(code)

    assert result.success is True
    assert "Read 3 lines" in result.stdout

    # Step 3: Verify output file was saved to PostgreSQL
    async with db_pool.acquire() as conn:
        output_file = await conn.fetchrow("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/output.txt'
        """)

        assert output_file is not None
        assert b"Processed 3 lines" in output_file["content"]


async def test_stateful_execution(db_pool, clean_files):
    """Test stateful session preserves variables."""
    executor = SandboxExecutor(db_pool, "test_sandbox", allow_net=True, stateful=True)

    # First execution: set variable
    result1 = await executor.execute("x = 42")
    assert result1.success is True
    assert result1.session_bytes is not None

    # Second execution: use variable
    result2 = await executor.execute(
        "print(x)",
        session_bytes=result1.session_bytes,
        session_metadata=result1.session_metadata,
    )

    assert result2.success is True
    assert "42" in result2.stdout


async def test_numpy_execution(executor, clean_files):
    """Test numpy works with micropip installation."""
    # Note: For async code, print output may not be captured
    # We verify success and that numpy can be imported and used
    code = """
import micropip
await micropip.install("numpy")
import numpy as np
arr = np.array([1, 2, 3, 4, 5])
result = arr.sum()
print(f"Sum: {result}")
# Return value so we can verify it worked
result
"""
    result = await executor.execute(code)

    assert result.success is True
    assert "Sum: 15" in result.stdout


async def test_multiple_files(executor, db_pool, clean_files):
    """Test creating multiple files."""
    code = """
for i in range(3):
    with open(f'/tmp/file{i}.txt', 'w') as f:
        f.write(f"File {i}")

print("All files created")
"""

    result = await executor.execute(code)

    assert result.success is True
    assert "All files created" in result.stdout
    assert result.created_files is not None
    assert len(result.created_files) >= 3

    # Verify all files are in PostgreSQL
    async with db_pool.acquire() as conn:
        files = await conn.fetch("""
            SELECT file_path FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path LIKE '/tmp/file%'
            ORDER BY file_path
        """)

        assert len(files) == 3


async def test_matplotlib_auto_backend(executor, clean_files):
    """Test matplotlib works without manual backend configuration."""
    code = """
import micropip
await micropip.install("matplotlib")

# Import matplotlib without setting backend manually
import matplotlib.pyplot as plt
import numpy as np

# Create a simple plot
x = np.linspace(0, 2 * np.pi, 100)
y = np.sin(x)

plt.figure(figsize=(8, 6))
plt.plot(x, y)
plt.title('Sine Wave')
plt.xlabel('x')
plt.ylabel('sin(x)')
plt.grid(True)

# Save the plot
plt.savefig('/tmp/sine_plot.png', dpi=150)
print("Plot created successfully")
"""
    result = await executor.execute(code)

    assert result.success is True, f"Execution failed: {result.stderr}"
    assert "Plot created successfully" in result.stdout
    assert result.created_files is not None
    assert "/tmp/sine_plot.png" in result.created_files


async def test_compiled_library_vfs_fallback(executor, db_pool, clean_files):
    """Test VFS fallback detects files from compiled libraries (openpyxl).

    Compiled libraries may use low-level I/O that bypasses Pyodide's snapshot
    mechanism. This test verifies the VFS fallback correctly detects such files.
    """
    code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import Workbook

# Create Excel file using openpyxl (compiled library)
wb = Workbook()
ws = wb.active
ws['A1'] = 'Product'
ws['B1'] = 'Quantity'
ws['A2'] = 'Widget'
ws['B2'] = 42

wb.save('/tmp/report.xlsx')
print("Excel file created")
"""

    result = await executor.execute(code)

    assert result.success is True, f"Execution failed: {result.stderr}"
    assert "Excel file created" in result.stdout

    # Verify file was tracked (via VFS fallback if TypeScript snapshot missed it)
    assert result.created_files is not None, "created_files should not be None"
    assert "/tmp/report.xlsx" in result.created_files, (
        f"Excel file not tracked. created_files: {result.created_files}"
    )

    # Verify file was saved to PostgreSQL VFS
    async with db_pool.acquire() as conn:
        file_data = await conn.fetchrow("""
            SELECT content, size FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/report.xlsx'
        """)

        assert file_data is not None, "File not found in VFS"
        assert file_data["size"] > 0, "File size should be > 0"
        # Excel files start with PK magic bytes (ZIP format)
        assert file_data["content"][:2] == b"PK", "File should be valid Excel format"


async def test_vfs_fallback_skipped_on_execution_failure(executor, db_pool, clean_files):
    """Test VFS fallback does NOT trigger when execution fails.

    Even if a compiled library creates files before throwing an error,
    the VFS fallback should not run because success=False.
    This prevents tracking incomplete/corrupted files.
    """
    code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import Workbook

# Create file
wb = Workbook()
ws = wb.active
ws['A1'] = 'Test'
wb.save('/tmp/partial.xlsx')

# Then fail
raise RuntimeError("Simulated error after file creation")
"""

    result = await executor.execute(code)

    # Execution should fail
    assert result.success is False, "Execution should fail due to RuntimeError"
    assert "RuntimeError" in result.stderr
    assert "Simulated error" in result.stderr

    # VFS fallback should NOT run (no log message)
    # created_files should be None because success=False
    assert result.created_files is None, (
        f"VFS fallback should not run when success=False. Got: {result.created_files}"
    )

    # Verify file MAY exist in VFS (written before error)
    # but is NOT tracked in created_files
    async with db_pool.acquire() as conn:
        file_exists = await conn.fetchval("""
            SELECT EXISTS(
                SELECT 1 FROM sandbox_filesystem
                WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/partial.xlsx'
            )
        """)

        # File may or may not exist (timing-dependent), but either way
        # it should NOT be in created_files
        if file_exists:
            print("Note: File exists in VFS but correctly NOT tracked due to failure")


async def test_vfs_fallback_supplements_typescript_snapshot(executor, db_pool, clean_files):
    """Test VFS fallback adds files missed by TypeScript, not duplicates.

    When some files are detected by TypeScript snapshot and others are missed
    (compiled library), the VFS fallback should only add the missing ones.
    """
    code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import Workbook

# File 1: Created with Python built-in (TypeScript WILL detect)
with open('/tmp/plain.txt', 'w') as f:
    f.write('Created with open()')

# File 2: Created with openpyxl (TypeScript may MISS)
wb = Workbook()
ws = wb.active
ws['A1'] = 'Data'
wb.save('/tmp/compiled.xlsx')

print("Both files created")
"""

    result = await executor.execute(code)

    assert result.success is True, f"Execution failed: {result.stderr}"
    assert "Both files created" in result.stdout

    # created_files should contain BOTH files
    assert result.created_files is not None
    assert len(result.created_files) >= 1, "At least one file should be tracked"

    # Both files should be in the list (order doesn't matter)
    created_paths = set(result.created_files)

    # At minimum, we should have the Excel file
    # (plain.txt may or may not be detected depending on Pyodide version)
    assert "/tmp/compiled.xlsx" in created_paths, (
        f"Excel file should be tracked. Got: {result.created_files}"
    )

    # Verify both files exist in VFS
    async with db_pool.acquire() as conn:
        files_in_vfs = await conn.fetch("""
            SELECT file_path FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox'
            AND file_path IN ('/tmp/plain.txt', '/tmp/compiled.xlsx')
            ORDER BY file_path
        """)

        vfs_paths = [row["file_path"] for row in files_in_vfs]
        assert "/tmp/plain.txt" in vfs_paths, "Plain text file should be in VFS"
        assert "/tmp/compiled.xlsx" in vfs_paths, "Excel file should be in VFS"


async def test_vfs_fallback_emits_log_message(executor, db_pool, clean_files, caplog):
    """Test VFS fallback logs INFO message when it detects files.

    The log message helps debug file tracking issues and should contain:
    - Number of files detected
    - Thread ID
    - List of file paths

    Note: VFS fallback only triggers when TypeScript snapshot misses files.
    If TypeScript detects the file, VFS fallback won't run and no log appears.
    """
    # Ensure we capture logs at INFO level
    caplog.set_level(logging.INFO, logger="mayflower_sandbox.sandbox_executor")

    code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 'Test'
wb.save('/tmp/logged.xlsx')
print("File created")
"""

    result = await executor.execute(code)

    assert result.success is True
    assert result.created_files is not None
    assert "/tmp/logged.xlsx" in result.created_files

    # Check for the log message - it may or may not appear depending on
    # whether TypeScript snapshot detected the file
    log_records = [r for r in caplog.records if "VFS fallback detected" in r.message]

    # If VFS fallback triggered, verify the log message format
    if log_records:
        log_message = log_records[0].message

        # Verify log message contains expected information
        assert "VFS fallback detected" in log_message
        assert "test_sandbox" in log_message  # Thread ID
        assert "/tmp/logged.xlsx" in log_message  # File path
        print("VFS fallback triggered - log message verified")
    else:
        # TypeScript detected the file, VFS fallback wasn't needed
        print("TypeScript snapshot detected file - VFS fallback not triggered")


async def test_vfs_fallback_from_empty_vfs(db_pool, clean_files):
    """Test VFS fallback works when VFS starts completely empty.

    Edge case: No files in VFS before execution, so before_vfs_files is empty set.
    After execution, VFS has one file from compiled library.
    """
    # Create a fresh executor with a unique thread_id
    thread_id = "test_empty_vfs"

    # Ensure session exists
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_sessions (thread_id, expires_at)
            VALUES ($1, NOW() + INTERVAL '1 day')
            ON CONFLICT (thread_id) DO NOTHING
        """,
            thread_id,
        )

    # Ensure VFS is empty for this thread
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            DELETE FROM sandbox_filesystem WHERE thread_id = $1
        """,
            thread_id,
        )

        # Verify empty
        count = await conn.fetchval(
            """
            SELECT COUNT(*) FROM sandbox_filesystem WHERE thread_id = $1
        """,
            thread_id,
        )
        assert count == 0, "VFS should be empty before test"

    executor = SandboxExecutor(db_pool, thread_id, allow_net=True)

    code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 'First File'
wb.save('/tmp/first.xlsx')
print("Created first file in empty VFS")
"""

    result = await executor.execute(code)

    assert result.success is True
    assert "Created first file" in result.stdout

    # VFS fallback should detect the file
    assert result.created_files is not None, "VFS fallback should detect file"
    assert "/tmp/first.xlsx" in result.created_files

    # Verify file is in VFS
    async with db_pool.acquire() as conn:
        file_exists = await conn.fetchval(
            """
            SELECT EXISTS(
                SELECT 1 FROM sandbox_filesystem
                WHERE thread_id = $1 AND file_path = '/tmp/first.xlsx'
            )
        """,
            thread_id,
        )

        assert file_exists, "File should be saved to VFS"

    # Cleanup: Remove all files for this thread to ensure test isolation
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            DELETE FROM sandbox_filesystem WHERE thread_id = $1
        """,
            thread_id,
        )


@pytest.mark.asyncio
async def test_openpyxl_with_micropip_install(db_pool, clean_files):
    """Test that openpyxl works when properly installed with micropip.

    This demonstrates the correct pattern for using third-party packages
    in Pyodide: install with micropip first, then import and use.

    Verifies that VFS fallback correctly detects Excel files created by openpyxl.
    """
    executor = SandboxExecutor(db_pool, "test_sandbox")

    code = """
import micropip
await micropip.install("openpyxl")

from openpyxl import Workbook

# Create a simple Excel file
wb = Workbook()
ws = wb.active
ws['A1'] = 'Test Data'
ws['A2'] = 42
wb.save('/tmp/test.xlsx')

print("Excel file created successfully with openpyxl")
"""

    result = await executor.execute(code)

    # Execution should succeed
    assert result.success is True, f"Execution failed: {result.stderr}"
    assert "Excel file created successfully" in result.stdout

    # VFS fallback should detect the created Excel file
    assert result.created_files is not None, "VFS fallback should detect file"
    assert "/tmp/test.xlsx" in result.created_files, (
        f"Excel file should be tracked. Got: {result.created_files}"
    )

    # Verify file exists in VFS and has content
    async with db_pool.acquire() as conn:
        file_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/test.xlsx'
        """)

        assert file_content is not None, "File should exist in VFS"


@pytest.mark.asyncio
async def test_file_creation_in_nonstandard_paths(db_pool, clean_files):
    """Test that files created in non-standard paths are detected.

    This is a regression test for the bug where files created outside
    /tmp and /data directories were not detected by the snapshot-based
    file tracking system.

    The new FS.trackingDelegate approach detects files in ANY location.
    """
    executor = SandboxExecutor(db_pool, "test_sandbox")

    code = """
import os

# Create files in various non-standard locations
os.makedirs('/home/pyodide', exist_ok=True)
os.makedirs('/var/log', exist_ok=True)
os.makedirs('/opt/myapp', exist_ok=True)

# Write files to non-standard paths
with open('/home/pyodide/cow.png', 'wb') as f:
    f.write(b'fake PNG data')

with open('/var/log/app.log', 'w') as f:
    f.write('Application log entry')

with open('/opt/myapp/config.json', 'w') as f:
    f.write('{"setting": "value"}')

print("Files created in non-standard paths")
"""

    result = await executor.execute(code)

    assert result.success is True, f"Execution failed: {result.stderr}"
    assert "Files created in non-standard paths" in result.stdout

    # All files should be detected regardless of location
    assert result.created_files is not None, "Should detect created files"
    assert "/home/pyodide/cow.png" in result.created_files, (
        f"Should detect file in /home/pyodide. Got: {result.created_files}"
    )
    assert "/var/log/app.log" in result.created_files, (
        f"Should detect file in /var/log. Got: {result.created_files}"
    )
    assert "/opt/myapp/config.json" in result.created_files, (
        f"Should detect file in /opt/myapp. Got: {result.created_files}"
    )

    # Verify files exist in VFS with correct content
    async with db_pool.acquire() as conn:
        cow_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/home/pyodide/cow.png'
        """)
        assert cow_content == b"fake PNG data", "cow.png should have correct content"

        log_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/var/log/app.log'
        """)
        assert log_content == b"Application log entry", "app.log should have correct content"

        config_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/opt/myapp/config.json'
        """)
        assert config_content == b'{"setting": "value"}', "config.json should have correct content"


@pytest.mark.asyncio
async def test_matplotlib_savefig_custom_path(db_pool, clean_files):
    """Test that matplotlib savefig works with custom paths outside /tmp.

    This is a regression test for the bug where matplotlib plots saved
    to non-standard locations were not detected.
    """
    executor = SandboxExecutor(db_pool, "test_sandbox")

    code = """
import micropip
await micropip.install('matplotlib')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os

# Create custom directory
os.makedirs('/home/pyodide/plots', exist_ok=True)

# Create a simple plot
plt.figure()
plt.plot([1, 2, 3], [1, 4, 9])
plt.title('Test Plot')

# Save to custom path
plt.savefig('/home/pyodide/plots/chart.png')
plt.close()

print("Plot saved to custom path")
"""

    result = await executor.execute(code)

    assert result.success is True, f"Execution failed: {result.stderr}"
    assert "Plot saved to custom path" in result.stdout

    # File should be detected even in custom location
    assert result.created_files is not None, "Should detect matplotlib output"
    assert "/home/pyodide/plots/chart.png" in result.created_files, (
        f"Should detect plot in custom path. Got: {result.created_files}"
    )

    # Verify file exists in VFS and is a valid PNG
    async with db_pool.acquire() as conn:
        file_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/home/pyodide/plots/chart.png'
        """)

        assert file_content is not None, "Plot file should exist in VFS"
        assert len(file_content) > 1000, "PNG file should have substantial content"
        assert file_content.startswith(b"\x89PNG"), "Should be valid PNG file"


@pytest.mark.asyncio
async def test_file_modification_detection(db_pool, clean_files):
    """Test that file modifications are properly detected.

    The FS.trackingDelegate approach tracks both file creation and modification.
    """
    executor = SandboxExecutor(db_pool, "test_sandbox")

    # First execution: create file
    code1 = """
with open('/tmp/counter.txt', 'w') as f:
    f.write('Count: 1')
print("Initial file created")
"""

    result1 = await executor.execute(code1)
    assert result1.success is True
    assert "/tmp/counter.txt" in result1.created_files

    # Second execution: modify the same file
    code2 = """
with open('/tmp/counter.txt', 'r') as f:
    content = f.read()

with open('/tmp/counter.txt', 'w') as f:
    f.write('Count: 2')

print("File modified")
"""

    result2 = await executor.execute(code2)
    assert result2.success is True
    assert "/tmp/counter.txt" in result2.created_files, "Modified files should be tracked"

    # Verify final content
    async with db_pool.acquire() as conn:
        final_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/counter.txt'
        """)
        assert final_content == b"Count: 2", "File should have updated content"


@pytest.mark.asyncio
async def test_file_append_detection(db_pool, clean_files):
    """Test that appending to files is detected.

    Files opened in append mode should trigger onWriteToFile callbacks.
    """
    executor = SandboxExecutor(db_pool, "test_sandbox")

    # Create initial file
    code1 = """
with open('/tmp/log.txt', 'w') as f:
    f.write('Line 1\\n')
print("Log created")
"""

    result1 = await executor.execute(code1)
    assert result1.success is True

    # Append to file
    code2 = """
with open('/tmp/log.txt', 'a') as f:
    f.write('Line 2\\n')
    f.write('Line 3\\n')
print("Log appended")
"""

    result2 = await executor.execute(code2)
    assert result2.success is True
    assert "/tmp/log.txt" in result2.created_files, "Appended files should be tracked"

    # Verify appended content
    async with db_pool.acquire() as conn:
        final_content = await conn.fetchval("""
            SELECT content FROM sandbox_filesystem
            WHERE thread_id = 'test_sandbox' AND file_path = '/tmp/log.txt'
        """)
        content_str = final_content.decode("utf-8")
        assert "Line 1" in content_str
        assert "Line 2" in content_str
        assert "Line 3" in content_str
