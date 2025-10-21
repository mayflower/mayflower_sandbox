"""
Tests for xlsx_helpers module using openpyxl.
"""

import os
import sys

import asyncpg
import pytest
from dotenv import load_dotenv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

load_dotenv()

from mayflower_sandbox.sandbox_executor import SandboxExecutor


@pytest.fixture
async def db_pool():
    """Create test database connection pool."""
    db_config = {
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "database": os.getenv("POSTGRES_DB", "mayflower_test"),
        "user": os.getenv("POSTGRES_USER", "postgres"),
        "password": os.getenv("POSTGRES_PASSWORD", "postgres"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
    }

    pool = await asyncpg.create_pool(**db_config)

    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO sandbox_sessions (thread_id, expires_at)
            VALUES ('xlsx_helpers_test', NOW() + INTERVAL '1 day')
            ON CONFLICT (thread_id) DO NOTHING
        """
        )

    yield pool
    await pool.close()


async def test_xlsx_import(db_pool):
    """Test that xlsx_helpers can be imported."""
    executor = SandboxExecutor(
        db_pool, "xlsx_helpers_test", allow_net=True, stateful=True, timeout_seconds=60.0
    )

    code = """
import micropip
await micropip.install('openpyxl')

from document.xlsx_helpers import (
    xlsx_get_sheet_names,
    xlsx_read_cells,
    xlsx_write_cells,
    xlsx_to_dict,
    xlsx_has_formulas,
    xlsx_read_with_formulas
)

print("✓ All xlsx_helpers functions imported successfully")
"""

    result = await executor.execute(code)
    assert result.success, f"Import failed: {result.stderr}"
    assert "✓" in result.stdout


async def test_xlsx_read_write(db_pool):
    """Test xlsx_read_cells and xlsx_write_cells."""
    executor = SandboxExecutor(
        db_pool, "xlsx_helpers_test", allow_net=True, stateful=True, timeout_seconds=60.0
    )

    code = """
import micropip
await micropip.install('openpyxl')

from openpyxl import Workbook
import io

# Create a test workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['A2'] = 'Alice'
ws['B2'] = 30

buf = io.BytesIO()
wb.save(buf)
xlsx_bytes = buf.getvalue()

# Test reading cells
from document.xlsx_helpers import xlsx_read_cells, xlsx_write_cells

values = xlsx_read_cells(xlsx_bytes, 'Sheet', ['A1', 'B1', 'A2', 'B2'])
print(f"Read values: {values}")

assert values['A1'] == 'Name', f"Expected 'Name', got {values['A1']}"
assert values['B1'] == 'Age', f"Expected 'Age', got {values['B1']}"
assert values['A2'] == 'Alice', f"Expected 'Alice', got {values['A2']}"
assert values['B2'] == 30, f"Expected 30, got {values['B2']}"

# Test writing cells
modified = xlsx_write_cells(xlsx_bytes, 'Sheet', {'A2': 'Bob', 'B2': 25})
new_values = xlsx_read_cells(modified, 'Sheet', ['A2', 'B2'])

assert new_values['A2'] == 'Bob', f"Expected 'Bob', got {new_values['A2']}"
assert new_values['B2'] == 25, f"Expected 25, got {new_values['B2']}"

print("✓ xlsx_read_cells and xlsx_write_cells work correctly")
"""

    result = await executor.execute(code)
    assert result.success, f"Test failed: {result.stderr}"
    assert "✓" in result.stdout


async def test_xlsx_to_dict(db_pool):
    """Test xlsx_to_dict function."""
    executor = SandboxExecutor(
        db_pool, "xlsx_helpers_test", allow_net=True, stateful=True, timeout_seconds=60.0
    )

    code = """
import micropip
await micropip.install('openpyxl')

from openpyxl import Workbook
import io

# Create a test workbook with headers
wb = Workbook()
ws = wb.active
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['A2'] = 'Alice'
ws['B2'] = 30
ws['A3'] = 'Bob'
ws['B3'] = 25

buf = io.BytesIO()
wb.save(buf)
xlsx_bytes = buf.getvalue()

from document.xlsx_helpers import xlsx_to_dict

data = xlsx_to_dict(xlsx_bytes, 'Sheet', has_header=True)
print(f"Converted data: {data}")

assert len(data) == 2, f"Expected 2 rows, got {len(data)}"
assert data[0]['Name'] == 'Alice', f"Expected 'Alice', got {data[0]['Name']}"
assert data[0]['Age'] == 30, f"Expected 30, got {data[0]['Age']}"
assert data[1]['Name'] == 'Bob', f"Expected 'Bob', got {data[1]['Name']}"
assert data[1]['Age'] == 25, f"Expected 25, got {data[1]['Age']}"

print("✓ xlsx_to_dict works correctly")
"""

    result = await executor.execute(code)
    assert result.success, f"Test failed: {result.stderr}"
    assert "✓" in result.stdout


async def test_xlsx_formulas(db_pool):
    """Test xlsx_has_formulas and xlsx_read_with_formulas."""
    executor = SandboxExecutor(
        db_pool, "xlsx_helpers_test", allow_net=True, stateful=True, timeout_seconds=60.0
    )

    code = """
import micropip
await micropip.install('openpyxl')

from openpyxl import Workbook
import io

# Create a test workbook with formulas
wb = Workbook()
ws = wb.active
ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = '=A1+A2'

buf = io.BytesIO()
wb.save(buf)
xlsx_bytes = buf.getvalue()

from document.xlsx_helpers import xlsx_has_formulas, xlsx_read_with_formulas

# Test has_formulas
formulas = xlsx_has_formulas(xlsx_bytes)
print(f"Formulas found: {formulas}")

assert 'Sheet' in formulas, "Expected to find 'Sheet' in formulas"
assert 'A3' in formulas['Sheet'], f"Expected 'A3' in formulas, got {formulas['Sheet']}"

# Test read_with_formulas
data = xlsx_read_with_formulas(xlsx_bytes)
print(f"Formula data: {data['formulas']}")

assert 'Sheet' in data['formulas'], "Expected 'Sheet' in formulas"
assert 'A3' in data['formulas']['Sheet'], "Expected 'A3' in formulas"
assert data['formulas']['Sheet']['A3'] == '=A1+A2', \
    f"Expected '=A1+A2', got {data['formulas']['Sheet']['A3']}"

print("✓ xlsx_has_formulas and xlsx_read_with_formulas work correctly")
"""

    result = await executor.execute(code)
    assert result.success, f"Test failed: {result.stderr}"
    assert "✓" in result.stdout
